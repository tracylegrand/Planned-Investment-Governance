#!/usr/bin/env python3
import openpyxl
import snowflake.connector
import os
import re

SPREADSHEET_PATH = '/Users/tlegrand/Downloads/Investments Tracker  - Majors .xlsx'
CONNECTION_NAME = os.getenv("SNOWFLAKE_CONNECTION_NAME") or "DemoAcct"

QUARTER_MAP = {
    'Q4 FY26': ('FY2026-Q4', '2025-11-01'),
    'Q3 FY26': ('FY2026-Q3', '2025-08-01'),
    'Q2 FY26': ('FY2026-Q2', '2025-05-01'),
    'Q1 FY26': ('FY2026-Q1', '2025-02-01'),
}

JB_APPROVAL_MAP = {
    'yes': 'FINAL_APPROVED',
    'y': 'FINAL_APPROVED',
    'no': 'REJECTED',
    'n': 'REJECTED',
    'maybe': 'SUBMITTED',
    'pending': 'SUBMITTED',
}

REGION_MAP = {
    'CME': 'Communications, Media & Entertainment',
    'FSI': 'Financial Services',
    'FSIGlobals': 'FSI Globals',
    'HCLS': 'Healthcare & Life Sciences',
    'MFG': 'Manufacturing',
    'RCG': 'Retail & Consumer Goods',
    'RetailCG': 'Retail & Consumer Goods'
}

SHEET_CONFIGS = {
    'Q4 FY26': {
        'industry_col': 1,       # A
        'priority_col': 2,       # B
        'customer_col': 3,       # C
        'scoped_col': 4,         # D - Project Scoped (y/n)
        'amount_col': 5,         # E
        'deal_structure_col': 6, # F
        'capacity_deal_col': 7,  # G - Attached to Capacity Deal
        'inv_type_col': 8,       # H
        'sfdc_link_col': 9,      # I
        'partner_col': 10,       # J
        'justification_col': 11, # K
        'jb_approval_col': 12,   # L
        'jb_comments_col': 13,   # M
        'sfdc_request_col': None,
    },
    'Q3 FY26': {
        'industry_col': 1,       # A
        'priority_col': 2,       # B
        'customer_col': 3,       # C
        'scoped_col': None,
        'amount_col': 4,         # D
        'deal_structure_col': 5, # E
        'capacity_deal_col': 6,  # F
        'inv_type_col': 7,       # G
        'sfdc_link_col': 8,      # H
        'partner_col': 9,        # I
        'justification_col': 10, # J
        'jb_approval_col': 11,   # K
        'jb_comments_col': 12,   # L
        'sfdc_request_col': None,
    },
    'Q2 FY26': {
        'industry_col': 1,       # A
        'priority_col': 2,       # B
        'customer_col': 3,       # C
        'scoped_col': None,
        'amount_col': 4,         # D
        'deal_structure_col': 5, # E
        'capacity_deal_col': 6,  # F
        'inv_type_col': 7,       # G
        'sfdc_link_col': 8,      # H
        'partner_col': 9,        # I
        'justification_col': 10, # J
        'jb_approval_col': 11,   # K
        'jb_comments_col': 12,   # L
        'sfdc_request_col': 17,  # Q - SPN ID
    },
    'Q1 FY26': {
        'industry_col': 1,       # A
        'priority_col': 2,       # B
        'customer_col': 3,       # C
        'scoped_col': None,
        'amount_col': 4,         # D
        'deal_structure_col': 5, # E
        'capacity_deal_col': 6,  # F
        'inv_type_col': 7,       # G
        'sfdc_link_col': 8,      # H
        'partner_col': 9,        # I
        'justification_col': 10, # J
        'jb_approval_col': 11,   # K
        'jb_comments_col': 14,   # N - JB Comment
        'sfdc_request_col': 18,  # R - SPN Record
    },
}

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, str) and val.upper() == 'TBD':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('nan', 'n/a', 'none') else None

def parse_boolean(val):
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('Y', 'YES', 'TRUE', '1'):
        return True
    elif s in ('N', 'NO', 'FALSE', '0'):
        return False
    return None

def extract_sfdc_link(cell):
    if cell.hyperlink:
        return cell.hyperlink.target
    text = safe_str(cell.value)
    if text:
        url_match = re.search(r'https?://[^\s]+', text)
        if url_match:
            return url_match.group(0)
    return None

def get_cell_value(ws, row, col):
    if col is None:
        return None
    return ws.cell(row=row, column=col).value

def map_jb_approval_to_status(val):
    if val is None:
        return 'DRAFT'
    s = str(val).strip().lower()
    return JB_APPROVAL_MAP.get(s, 'DRAFT')

def load_investments():
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    
    conn = snowflake.connector.connect(connection_name=CONNECTION_NAME)
    cursor = conn.cursor()
    
    cursor.execute("USE DATABASE TEMP")
    cursor.execute("USE SCHEMA INVESTMENT_GOVERNANCE")
    
    cursor.execute("""
        SELECT ACCOUNT_NAME, REP_NAME, AE_EMPLOYEE_ID 
        FROM SALES.ACCOUNT_BASIC.DETAILS 
        WHERE REP_NAME IS NOT NULL
    """)
    ae_lookup = {}
    for row in cursor.fetchall():
        ae_lookup[row[0].upper() if row[0] else ''] = {'rep_name': row[1], 'employee_id': row[2]}
    print(f"Loaded {len(ae_lookup)} account AE mappings from SFDC")
    
    total_inserted = 0
    total_skipped = 0
    
    for sheet_name in ['Q4 FY26', 'Q3 FY26', 'Q2 FY26', 'Q1 FY26']:
        ws = wb[sheet_name]
        config = SHEET_CONFIGS[sheet_name]
        quarter_info = QUARTER_MAP[sheet_name]
        investment_quarter = quarter_info[0]
        created_at = quarter_info[1]
        
        print(f"\n{'='*60}")
        print(f"Processing: {sheet_name} -> {investment_quarter}")
        print(f"{'='*60}")
        
        inserted = 0
        skipped = 0
        
        for row_num in range(6, ws.max_row + 1):
            customer = get_cell_value(ws, row_num, config['customer_col'])
            amount = get_cell_value(ws, row_num, config['amount_col'])
            
            if not customer or not str(customer).strip():
                continue
            
            customer = str(customer).strip()
            amount_val = safe_float(amount)
            
            if amount_val is None or amount_val == 0:
                skipped += 1
                continue
            
            industry = safe_str(get_cell_value(ws, row_num, config['industry_col']))
            priority = safe_str(get_cell_value(ws, row_num, config['priority_col']))
            deal_structure = safe_str(get_cell_value(ws, row_num, config['deal_structure_col']))
            inv_type = safe_str(get_cell_value(ws, row_num, config['inv_type_col']))
            partner = safe_str(get_cell_value(ws, row_num, config['partner_col']))
            justification = safe_str(get_cell_value(ws, row_num, config['justification_col']))
            jb_approval = get_cell_value(ws, row_num, config['jb_approval_col'])
            jb_comments = safe_str(get_cell_value(ws, row_num, config['jb_comments_col']))
            
            scoped = None
            if config['scoped_col']:
                scoped = parse_boolean(get_cell_value(ws, row_num, config['scoped_col']))
            
            capacity_deal = None
            if config['capacity_deal_col']:
                capacity_deal = parse_boolean(get_cell_value(ws, row_num, config['capacity_deal_col']))
            
            sfdc_link_cell = ws.cell(row=row_num, column=config['sfdc_link_col']) if config['sfdc_link_col'] else None
            sfdc_link = extract_sfdc_link(sfdc_link_cell) if sfdc_link_cell else None
            
            sfdc_request_id = None
            if config['sfdc_request_col']:
                sfdc_request_id = safe_str(get_cell_value(ws, row_num, config['sfdc_request_col']))
            
            industry_segment = REGION_MAP.get(industry, industry if industry else 'Enterprise')
            status = map_jb_approval_to_status(jb_approval)
            
            if status == 'FINAL_APPROVED':
                approval_level = 5
            elif status == 'SUBMITTED':
                approval_level = 1
            elif status == 'REJECTED':
                approval_level = 0
            else:
                approval_level = 0
            
            if deal_structure:
                request_title = f"{customer} - {deal_structure}"
            else:
                request_title = f"{customer} - {investment_quarter} Investment"
            
            # Truncate request_title to fit database column (500 chars)
            if len(request_title) > 500:
                request_title = request_title[:497] + "..."
            
            if not inv_type or inv_type in ('Prioritized Accounts', 'Strategic'):
                inv_type = 'PS&T'
            elif 'AMP' in inv_type.upper():
                inv_type = 'AMP'
            elif 'PS' in inv_type.upper():
                inv_type = 'PS&T'
            
            ae_info = ae_lookup.get(customer.upper(), {})
            created_by = ae_info.get('rep_name', 'tlegrand')
            created_by_name = ae_info.get('rep_name', 'Todd Legrand')
            created_by_employee_id = ae_info.get('employee_id')
            
            insert_sql = """
            INSERT INTO INVESTMENT_REQUESTS (
                REQUEST_TITLE,
                ACCOUNT_NAME,
                INVESTMENT_TYPE,
                REQUESTED_AMOUNT,
                INVESTMENT_QUARTER,
                BUSINESS_JUSTIFICATION,
                THEATER,
                INDUSTRY_SEGMENT,
                STATUS,
                CURRENT_APPROVAL_LEVEL,
                CREATED_BY,
                CREATED_BY_NAME,
                CREATED_BY_EMPLOYEE_ID,
                CREATED_AT,
                PARTNER_ATTACHED,
                IN_SALESFORCE,
                GVP_COMMENTS,
                SCOPED,
                OPPORTUNITY_NAME,
                SFDC_OPPORTUNITY_LINK,
                SFDC_REQUEST_ID,
                PRIORITY
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            cursor.execute(insert_sql, (
                request_title,
                customer,
                inv_type,
                amount_val,
                investment_quarter,
                justification,
                'USMajors',
                industry_segment,
                status,
                approval_level,
                created_by,
                created_by_name,
                created_by_employee_id,
                created_at,
                partner,
                capacity_deal,
                jb_comments,
                scoped,
                deal_structure,
                sfdc_link,
                sfdc_request_id,
                priority
            ))
            inserted += 1
            print(f"  {customer} | ${amount_val:,.0f} | {status} | {priority or 'N/A'}")
        
        print(f"\n  Sheet {sheet_name}: Inserted {inserted}, Skipped {skipped}")
        total_inserted += inserted
        total_skipped += skipped
    
    conn.commit()
    
    cursor.execute("""
        SELECT INVESTMENT_QUARTER, COUNT(*) as CNT, SUM(REQUESTED_AMOUNT) as TOTAL
        FROM INVESTMENT_REQUESTS
        GROUP BY INVESTMENT_QUARTER
        ORDER BY INVESTMENT_QUARTER
    """)
    
    print(f"\n{'='*60}")
    print("Results by Quarter (All Data):")
    print(f"{'='*60}")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]} records, ${row[2]:,.0f}")
    
    cursor.execute("SELECT COUNT(*) FROM INVESTMENT_REQUESTS")
    final_count = cursor.fetchone()[0]
    
    print(f"\nFY26 Import Summary:")
    print(f"  Inserted: {total_inserted}")
    print(f"  Skipped: {total_skipped}")
    print(f"  Total records in DB: {final_count}")
    
    cursor.close()
    conn.close()
    
    return total_inserted

if __name__ == "__main__":
    load_investments()
