#!/usr/bin/env python3
"""Refresh budget data from a Maher Copy spreadsheet."""

import csv
import os
import sys
from datetime import date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
BUDGET_DIR = os.path.join(PROJECT_DIR, "budget_data")
CSV_PATH = os.path.join(BUDGET_DIR, "budget_data.csv")
HTML_PATH = os.path.join(BUDGET_DIR, "Source Financial Data.html")

LABEL_MAP = {
    "CME": "CME (TMT)",
    "FSI": "FSI",
    "FSIGlobals": "FSIGlobals",
    "HCLS": "HCLS",
    "MFG": "MFG",
    "RetailCG": "RCG",
}

FY27_ROWS = range(10, 16)
FY27_COLS = (28, 29, 30, 31, 32)
FY26_ROWS = range(5, 11)
FY26_COLS = (20, 21, 22, 23, 24)


def extract_budgets(wb, sheet_name, rows, cols, fiscal_year, label_col=2):
    ws = wb[sheet_name]
    budgets = []
    for row_num in rows:
        orig = ws.cell(row=row_num, column=label_col if fiscal_year == "FY2027" else 1).value or ""
        label = LABEL_MAP.get(orig.strip(), orig.strip())
        q1 = round(ws.cell(row=row_num, column=cols[0]).value or 0)
        q2 = round(ws.cell(row=row_num, column=cols[1]).value or 0)
        q3 = round(ws.cell(row=row_num, column=cols[2]).value or 0)
        q4 = round(ws.cell(row=row_num, column=cols[3]).value or 0)
        fy = round(ws.cell(row=row_num, column=cols[4]).value or 0)
        budgets.append({
            "fiscal_year": fiscal_year,
            "portfolio": label,
            "q1_budget": q1,
            "q2_budget": q2,
            "q3_budget": q3,
            "q4_budget": q4,
            "budget_amount": fy,
            "theater": "US Majors",
            "orig_label": orig,
        })
    return budgets


def write_csv(all_budgets):
    os.makedirs(BUDGET_DIR, exist_ok=True)
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "fiscal_year", "portfolio", "q1_budget", "q2_budget",
            "q3_budget", "q4_budget", "budget_amount", "theater",
        ])
        writer.writeheader()
        for b in all_budgets:
            row = {k: v for k, v in b.items() if k != "orig_label"}
            writer.writerow(row)
    print(f"  Wrote {len(all_budgets)} rows to {CSV_PATH}")


def fmt(v):
    return f"${v:,.0f}"


def write_html(fy27, fy26, input_file):
    def total_row(budgets):
        return {
            "q1": sum(b["q1_budget"] for b in budgets),
            "q2": sum(b["q2_budget"] for b in budgets),
            "q3": sum(b["q3_budget"] for b in budgets),
            "q4": sum(b["q4_budget"] for b in budgets),
            "fy": sum(b["budget_amount"] for b in budgets),
        }

    def table_rows(budgets):
        lines = []
        for b in budgets:
            lines.append(f'<tr><td>{b["portfolio"]}</td><td>{fmt(b["q1_budget"])}</td><td>{fmt(b["q2_budget"])}</td><td>{fmt(b["q3_budget"])}</td><td>{fmt(b["q4_budget"])}</td><td>{fmt(b["budget_amount"])}</td></tr>')
        t = total_row(budgets)
        lines.append(f'<tr class="total-row"><td>US Majors Total</td><td>{fmt(t["q1"])}</td><td>{fmt(t["q2"])}</td><td>{fmt(t["q3"])}</td><td>{fmt(t["q4"])}</td><td>{fmt(t["fy"])}</td></tr>')
        return "\n".join(lines)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Source Financial Data — Planned Investment Governance</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 960px; margin: 40px auto; padding: 0 20px; color: #333; }}
  h1 {{ border-bottom: 2px solid #2196f3; padding-bottom: 8px; }}
  h2 {{ color: #1565c0; margin-top: 32px; }}
  table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
  th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: right; }}
  th {{ background: #f5f5f5; font-weight: 600; }}
  td:first-child, th:first-child {{ text-align: left; }}
  tr:nth-child(even) {{ background: #fafafa; }}
  .total-row {{ font-weight: bold; background: #e3f2fd !important; }}
  .meta {{ color: #666; font-size: 0.9em; margin: 4px 0; }}
  .mapping {{ background: #fff3e0; border: 1px solid #ffe0b2; border-radius: 4px; padding: 12px 16px; margin: 16px 0; }}
  .mapping dt {{ font-weight: 600; }}
  .mapping dd {{ margin: 0 0 8px 16px; }}
</style>
</head>
<body>
<h1>Source Financial Data</h1>
<p class="meta">Extracted from: <strong>{os.path.basename(input_file)}</strong></p>
<p class="meta">Extraction date: {date.today().strftime("%B %d, %Y")}</p>

<div class="mapping">
<h3>Label Mapping (Maher Copy &rarr; Application)</h3>
<dl>
  <dt>CME &rarr; CME (TMT)</dt><dd>Consumer, Media &amp; Entertainment / TMT.</dd>
  <dt>FSI &rarr; FSI</dt><dd>Financial Services &amp; Insurance.</dd>
  <dt>FSIGlobals &rarr; FSIGlobals</dt><dd>FSI Global accounts.</dd>
  <dt>HCLS &rarr; HCLS</dt><dd>Healthcare &amp; Life Sciences.</dd>
  <dt>MFG &rarr; MFG</dt><dd>Manufacturing &amp; Industrial.</dd>
  <dt>RetailCG &rarr; RCG</dt><dd>Retail &amp; Consumer Goods.</dd>
</dl>
</div>

<h2>FY2027 — Investment Budget (Plan)</h2>
<table>
<thead><tr><th>Portfolio</th><th>Q1</th><th>Q2</th><th>Q3</th><th>Q4</th><th>FY Total</th></tr></thead>
<tbody>
{table_rows(fy27)}
</tbody>
</table>

<h2>FY2026 — Investment Actuals</h2>
<table>
<thead><tr><th>Portfolio</th><th>Q1</th><th>Q2</th><th>Q3</th><th>Q4</th><th>FY Total</th></tr></thead>
<tbody>
{table_rows(fy26)}
</tbody>
</table>

<h2>Data Source Notes</h2>
<ul>
<li>Editable budget data is maintained in <code>budget_data/budget_data.csv</code>.</li>
<li>Use <code>scripts/refresh_budgets.py</code> to import from a new Maher Copy file.</li>
</ul>
</body>
</html>"""
    with open(HTML_PATH, "w") as f:
        f.write(html)
    print(f"  Wrote HTML documentation to {HTML_PATH}")


def trigger_import(all_budgets):
    import requests
    api_url = os.getenv("PIG_API_URL", "http://localhost:8770")
    payload = {"budgets": [
        {k: v for k, v in b.items() if k != "orig_label"}
        for b in all_budgets
    ]}
    resp = requests.post(f"{api_url}/api/budgets/import", json=payload, timeout=30)
    resp.raise_for_status()
    result = resp.json()
    print(f"  Imported {result.get('imported', 0)} rows into database.")


def main():
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    print("Budget Data Refresh Tool")
    print("=" * 40)

    input_file = input("Enter path to Maher Copy spreadsheet: ").strip()
    if not input_file:
        print("No file specified. Exiting.")
        sys.exit(1)
    input_file = os.path.expanduser(input_file)
    if not os.path.isfile(input_file):
        print(f"File not found: {input_file}")
        sys.exit(1)

    print(f"\nLoading {input_file}...")
    wb = openpyxl.load_workbook(input_file, data_only=True)

    print("Extracting FY2027 budgets from 'FY27 Bookings Targets'...")
    fy27 = extract_budgets(wb, "FY27 Bookings Targets", FY27_ROWS, FY27_COLS, "FY2027")
    for b in fy27:
        print(f"  {b['portfolio']:15s}  Q1={fmt(b['q1_budget'])}  Q2={fmt(b['q2_budget'])}  Q3={fmt(b['q3_budget'])}  Q4={fmt(b['q4_budget'])}  FY={fmt(b['budget_amount'])}")

    print("\nExtracting FY2026 actuals from 'FY26 Bookings Historicals'...")
    fy26 = extract_budgets(wb, "FY26 Bookings Historicals", FY26_ROWS, FY26_COLS, "FY2026", label_col=1)
    for b in fy26:
        print(f"  {b['portfolio']:15s}  Q1={fmt(b['q1_budget'])}  Q2={fmt(b['q2_budget'])}  Q3={fmt(b['q3_budget'])}  Q4={fmt(b['q4_budget'])}  FY={fmt(b['budget_amount'])}")

    all_budgets = fy27 + fy26

    print(f"\nWriting CSV...")
    write_csv(all_budgets)

    print("Writing HTML documentation...")
    write_html(fy27, fy26, input_file)

    answer = input("\nImport into application database? (y/n): ").strip().lower()
    if answer == "y":
        print("Importing...")
        try:
            trigger_import(all_budgets)
            print("Done!")
        except Exception as e:
            print(f"Import failed: {e}")
            print("You can import later via the Financials admin panel.")
    else:
        print("Skipped database import. You can import later via the Financials admin panel.")


if __name__ == "__main__":
    main()
