#!/usr/bin/env python3
import openpyxl
import snowflake.connector
import os
import re
from datetime import datetime
from difflib import SequenceMatcher

SPREADSHEET_PATH = '/Users/tlegrand/Downloads/USMajors - FY27 Investment Tracker.xlsx'
CONNECTION_NAME = os.getenv("SNOWFLAKE_CONNECTION_NAME") or "DemoAcct"

QUARTER_MAP = {
    'Q126': ('FY2026-Q1', '2025-02-01'),
    'Q226': ('FY2026-Q2', '2025-05-01'),
    'Q326': ('FY2026-Q3', '2025-08-01'),
    'Q426': ('FY2026-Q4', '2025-11-01'),
    'Q127': ('FY2027-Q1', '2026-02-01'),
    'Q227': ('FY2027-Q2', '2026-05-01'),
    'Q327': ('FY2027-Q3', '2026-08-01'),
    'Q427': ('FY2027-Q4', '2026-11-01'),
}

STATUS_MAP = {
    'Approved': 'FINAL_APPROVED',
    'Approval Pending': 'SUBMITTED',
    'Not Approved': 'REJECTED',
}

REGION_MAP = {
    'CME': 'SCE',
    'FSI': 'FSI',
    'FSIGlobals': 'FSIGlobals',
    'HCLS': 'HCLS',
    'MFG': 'MFG',
    'RCG': 'RCG',
    'RetailCG': 'RCG',
    'SCE': 'SCE'
}

def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != 'nan' else None

def parse_boolean(val):
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('Y', 'YES', 'TRUE', '1'):
        return True
    elif s in ('N', 'NO', 'FALSE', '0'):
        return False
    return None

def extract_sfdc_link(cell):
    """Extract hyperlink URL from cell if present, otherwise check for URL in text"""
    if cell.hyperlink:
        return cell.hyperlink.target
    text = safe_str(cell.value)
    if text:
        url_match = re.search(r'https?://[^\s]+', text)
        if url_match:
            return url_match.group(0)
    return None

def map_status(approval_status, amount_val):
    """Map approval status to database status, defaulting based on amount"""
    if approval_status and str(approval_status).strip():
        status_str = str(approval_status).strip()
        if status_str in STATUS_MAP:
            return STATUS_MAP[status_str]
    
    # Empty approval status: approved if amount > 0, denied if amount <= 0
    if amount_val and amount_val > 0:
        return 'FINAL_APPROVED'
    return 'DENIED'

def fuzzy_match_account(customer_name, ae_lookup, threshold=0.85):
    """Find the best matching account name using fuzzy matching."""
    customer_upper = customer_name.upper()
    
    # Try exact match first
    if customer_upper in ae_lookup:
        return ae_lookup[customer_upper]
    
    # Try fuzzy matching
    best_match = None
    best_ratio = 0
    
    for account_name, ae_info in ae_lookup.items():
        ratio = SequenceMatcher(None, customer_upper, account_name).ratio()
        if ratio > best_ratio and ratio >= threshold:
            best_ratio = ratio
            best_match = ae_info
    
    return best_match

def load_investments():
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws = wb['Investment Details <<Input Here']
    
    conn = snowflake.connector.connect(connection_name=CONNECTION_NAME)
    cursor = conn.cursor()
    
    cursor.execute("USE DATABASE TEMP")
    cursor.execute("USE SCHEMA INVESTMENT_GOVERNANCE")
    
    cursor.execute("""
        SELECT ACCOUNT_NAME, REP_NAME, AE_EMPLOYEE_ID 
        FROM SALES.ACCOUNT_BASIC.DETAILS 
        WHERE REP_NAME IS NOT NULL
    """)
    ae_lookup = {}
    for row in cursor.fetchall():
        ae_lookup[row[0].upper() if row[0] else ''] = {'rep_name': row[1], 'employee_id': row[2]}
    print(f"Loaded {len(ae_lookup)} account AE mappings from SFDC")
    
    inserted = 0
    skipped = 0
    
    for row_num in range(10, ws.max_row + 1):
        # Column mappings:
        # A(1)=empty, B(2)=Funding Qtr, C(3)=Region, D(4)=Customer
        # E(5)=Opportunity Name + SFDC Link, F(6)=OppTACV, G(7)=OppTCV
        # H(8)=Invest Type, I(9)=Invest $, J(10)=Justification
        # K(11)=Partner, L(12)=Approval Status, M(13)=In SFDC?
        # N(14)=Scoped?, O(15)=Use Case, P(16)=Co-Invest
        # Q(17)=JB Comment, R(18)=AE Response, S(19)=SFDC Request #
        
        funding_qtr = ws.cell(row=row_num, column=2).value
        region = ws.cell(row=row_num, column=3).value
        customer = ws.cell(row=row_num, column=4).value
        opp_cell = ws.cell(row=row_num, column=5)
        opportunity_name = safe_str(opp_cell.value)
        sfdc_opp_link = extract_sfdc_link(opp_cell)
        opp_tacv = ws.cell(row=row_num, column=6).value
        opp_tcv = ws.cell(row=row_num, column=7).value
        investment_type = ws.cell(row=row_num, column=8).value
        invest_amt = ws.cell(row=row_num, column=9).value
        justification = ws.cell(row=row_num, column=10).value
        partner = ws.cell(row=row_num, column=11).value
        approval_status = ws.cell(row=row_num, column=12).value
        in_sfdc = ws.cell(row=row_num, column=13).value
        scoped = ws.cell(row=row_num, column=14).value
        use_cases = ws.cell(row=row_num, column=15).value
        co_invest = ws.cell(row=row_num, column=16).value
        jb_comment = ws.cell(row=row_num, column=17).value
        ae_response = ws.cell(row=row_num, column=18).value
        sfdc_request_id = ws.cell(row=row_num, column=19).value
        
        if not customer or not invest_amt:
            skipped += 1
            continue
        
        customer = str(customer).strip()
        amount_val = safe_float(invest_amt)
        
        quarter_info = QUARTER_MAP.get(funding_qtr, ('FY2027-Q1', '2026-02-01'))
        investment_quarter = quarter_info[0]
        created_at = quarter_info[1]
        
        status = map_status(approval_status, amount_val)
        
        if status == 'FINAL_APPROVED':
            approval_level = 5
        elif status == 'SUBMITTED':
            approval_level = 1
        elif status in ('REJECTED', 'DENIED'):
            approval_level = 0
        else:
            approval_level = 0
        
        inv_type = safe_str(investment_type) or 'PS&T'
        industry_segment = REGION_MAP.get(region, region if region else 'Enterprise')
        
        # Request title from opportunity name or customer
        if opportunity_name:
            request_title = f"{customer} - {opportunity_name}"
        else:
            request_title = f"{customer} - {investment_quarter} Investment"
        
        # Truncate request_title to fit database column (500 chars)
        if len(request_title) > 500:
            request_title = request_title[:497] + "..."
        
        ae_info = fuzzy_match_account(customer, ae_lookup)
        if ae_info:
            created_by = ae_info.get('rep_name', 'system')
            created_by_name = ae_info.get('rep_name', 'System Generated')
            created_by_employee_id = ae_info.get('employee_id')
        else:
            created_by = 'system'
            created_by_name = 'System Generated'
            created_by_employee_id = None
        
        insert_sql = """
        INSERT INTO INVESTMENT_REQUESTS (
            REQUEST_TITLE,
            ACCOUNT_NAME,
            INVESTMENT_TYPE,
            REQUESTED_AMOUNT,
            INVESTMENT_QUARTER,
            BUSINESS_JUSTIFICATION,
            THEATER,
            INDUSTRY_SEGMENT,
            STATUS,
            CURRENT_APPROVAL_LEVEL,
            CREATED_BY,
            CREATED_BY_NAME,
            CREATED_BY_EMPLOYEE_ID,
            CREATED_AT,
            OPPORTUNITY_TACV,
            OPPORTUNITY_TCV,
            PARTNER_ATTACHED,
            IN_SALESFORCE,
            ASSOCIATED_USE_CASES,
            CO_INVESTMENT,
            GVP_COMMENTS,
            SCOPED,
            OPPORTUNITY_NAME,
            SFDC_OPPORTUNITY_LINK,
            RESUBMISSION_COMMENT,
            SFDC_REQUEST_ID
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(insert_sql, (
            request_title,
            customer,
            inv_type,
            amount_val,
            investment_quarter,
            safe_str(justification),
            'US Majors',
            industry_segment,
            status,
            approval_level,
            created_by,
            created_by_name,
            created_by_employee_id,
            created_at,
            safe_float(opp_tacv),
            safe_float(opp_tcv),
            safe_str(partner),
            parse_boolean(in_sfdc),
            safe_str(use_cases),
            safe_str(co_invest),
            safe_str(jb_comment),
            parse_boolean(scoped),
            opportunity_name,
            sfdc_opp_link,
            safe_str(ae_response),
            safe_str(sfdc_request_id)
        ))
        inserted += 1
        print(f"  Inserted: {customer} | {investment_quarter} | ${invest_amt} | {status}")
    
    conn.commit()
    
    cursor.execute("""
        SELECT INVESTMENT_QUARTER, COUNT(*) as CNT, SUM(REQUESTED_AMOUNT) as TOTAL
        FROM INVESTMENT_REQUESTS
        GROUP BY INVESTMENT_QUARTER
        ORDER BY INVESTMENT_QUARTER
    """)
    
    print(f"\nResults by Quarter:")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]} records, ${row[2]:,.0f}")
    
    cursor.execute("SELECT COUNT(*) FROM INVESTMENT_REQUESTS")
    final_count = cursor.fetchone()[0]
    
    print(f"\nSummary:")
    print(f"  Inserted: {inserted}")
    print(f"  Skipped (empty): {skipped}")
    print(f"  Total records: {final_count}")
    
    cursor.close()
    conn.close()
    
    return inserted

if __name__ == "__main__":
    load_investments()
